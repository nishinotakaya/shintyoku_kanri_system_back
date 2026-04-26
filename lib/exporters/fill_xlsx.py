#!/usr/bin/env python3
"""既存 xlsx テンプレートを書式保持したままセル値だけ差し替える。
Rails の Exporter から呼ばれる。

usage: fill_xlsx.py <template> <output> <json_payload>

json_payload:
  {
    "sheet": 0,
    "cells": [
      {"row": 4, "col": 3, "value": "西野 鷹也"},        # 1-indexed (Excel座標)
      {"row": 7, "col": 1, "value": "2026-03-01", "type": "date"},
      ...
    ]
  }
"""
import sys
import json
from datetime import datetime
import openpyxl


def parse_value(spec):
    v = spec.get("value")
    t = spec.get("type")
    if v is None:
        return None
    if t == "date" and isinstance(v, str):
        return datetime.fromisoformat(v).date()
    return v


def main():
    template = sys.argv[1]
    output = sys.argv[2]
    payload = json.loads(sys.argv[3])

    wb = openpyxl.load_workbook(template)
    sheet_index = payload.get("sheet", 0)
    ws = wb[wb.sheetnames[sheet_index]]

    # シート名リネーム
    new_name = payload.get("sheet_name")
    if new_name:
        ws.title = new_name

    # A1 セル (日付ヘッダ) の年月も更新
    header_date = payload.get("header_date")
    if header_date:
        ws.cell(row=1, column=1).value = datetime.fromisoformat(header_date).date()

    for spec in payload.get("cells", []):
        row = spec["row"]
        col = spec["col"]
        value = parse_value(spec)
        cell = ws.cell(row=row, column=col)
        cell.value = value

    wb.save(output)
    print(output)


if __name__ == "__main__":
    main()
