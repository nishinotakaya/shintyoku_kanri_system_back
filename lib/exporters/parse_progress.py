#!/usr/bin/env python3
"""進捗管理 Excel を読み取り、タスク一覧を JSON で出力する。
列構成を自動検出するので、複数フォーマットに対応。

usage: parse_progress.py <file.xlsx>
"""
import sys, json, re
from datetime import datetime
import openpyxl


def parse_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        v = v.strip().strip("'")
        # "2026-04-06" 形式
        m = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", v)
        if m:
            return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
        # "11月4日" 形式
        m = re.search(r"(\d{1,2})月(\d{1,2})日", v)
        if m:
            return f"2025-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    return None


def short_name(title, sap):
    """タスクの短縮名を生成"""
    if sap:
        return sap
    t = re.sub(r'^[・\s]+', '', title)
    t = t.split('\n')[0].strip()
    t = re.sub(r'SAP-\d+\s*', '', t, flags=re.IGNORECASE).strip()
    t = re.sub(r'[（(][^)）]*[)）]', '', t).strip()
    t = re.sub(r'^[・→\s]+', '', t)
    t = re.sub(r'(について|に関して|の件|の対応|の改修|の修正|の調査|すべて|のため)$', '', t).strip()
    t = re.sub(r'[\s　]+', '', t)
    if len(t) > 15:
        t = t[:15]
    return t or title[:15]


def detect_layout(ws):
    """列構成を自動検出。
    タイトル列 = 最初のテキストが多い列 (A or B)
    日付列 = タイトル列の右側で datetime が入っている列群
    """
    # row 1-10 をスキャンして A列 or B列 どちらがタイトルか判定
    a_texts = sum(1 for r in range(1, 11) if isinstance(ws.cell(row=r, column=1).value, str))
    b_texts = sum(1 for r in range(1, 11) if isinstance(ws.cell(row=r, column=2).value, str))

    if b_texts > a_texts:
        # 進捗管理_西野.xlsx 形式 (B列タイトル, F-I列日付)
        return {"title": 2, "plan_start": 6, "plan_end": 7,
                "actual_start": 8, "actual_end": 9, "progress": 10, "start_row": 5}
    else:
        # 作業 勤怠に投げる用.xlsx 形式 (A列タイトル, E-H列日付)
        return {"title": 1, "plan_start": 5, "plan_end": 6,
                "actual_start": 7, "actual_end": 8, "progress": 9, "start_row": 1}


def main():
    path = sys.argv[1]
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    layout = detect_layout(ws)
    tasks = []

    for r in range(layout["start_row"], ws.max_row + 1):
        title_raw = ws.cell(row=r, column=layout["title"]).value
        if not title_raw or not isinstance(title_raw, str):
            continue
        title = title_raw.strip().strip("'")
        if not title or title.startswith("【") or title.startswith("["):
            continue

        sap_match = re.search(r"(SAP-\d+)", title, re.IGNORECASE)
        sap = sap_match.group(1).upper() if sap_match else None

        # 実績日
        actual_start = parse_date(ws.cell(row=r, column=layout["actual_start"]).value)
        actual_end = parse_date(ws.cell(row=r, column=layout["actual_end"]).value)

        # 予定日 fallback
        if not actual_start:
            actual_start = parse_date(ws.cell(row=r, column=layout["plan_start"]).value)
        if not actual_end:
            actual_end = parse_date(ws.cell(row=r, column=layout["plan_end"]).value)

        # 進捗率
        prog_raw = ws.cell(row=r, column=layout["progress"]).value
        progress = None
        if isinstance(prog_raw, (int, float)):
            progress = float(prog_raw)
        elif isinstance(prog_raw, str):
            m = re.search(r"([\d.]+)%", prog_raw)
            if m:
                progress = float(m.group(1)) / 100.0

        if actual_start:
            tasks.append({
                "title": title[:80],
                "sap": sap,
                "short": short_name(title, sap),
                "actual_start": actual_start,
                "actual_end": actual_end or actual_start,
                "progress": progress
            })

    print(json.dumps(tasks, ensure_ascii=False))


if __name__ == "__main__":
    main()
